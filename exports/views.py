import io
from django.http import FileResponse, HttpResponse
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from drf_spectacular.utils import extend_schema
from django.utils import timezone

from orders.models import Order
from accounts.views import IsAdminRole

# ReportLab untuk PDF
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

# OpenPyXL untuk XLSX
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

class ExportOrdersPDFView(APIView):
    """
    Endpoint untuk mengekspor data pesanan ke PDF.
    Hanya bisa diakses oleh role admin.
    """
    permission_classes = [IsAdminRole]

    @extend_schema(
        summary='Export Pesanan (PDF)',
        description='Mengunduh daftar pesanan dalam format PDF. Hanya untuk admin.',
        responses={200: bytes}
    )
    def get(self, request):
        orders = Order.objects.all().select_related('user', 'ticket__seminar')

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
        elements = []

        # Judul
        styles = getSampleStyleSheet()
        title_style = styles['Heading1']
        title_style.alignment = 1 # Center
        elements.append(Paragraph('Laporan Pesanan Tiket Seminar', title_style))
        elements.append(Spacer(1, 20))

        # Tabel Data
        data = [['No', 'Order ID', 'User', 'Seminar', 'Tipe Tiket', 'Qty', 'Total Harga', 'Status', 'Tanggal']]
        
        for idx, order in enumerate(orders, 1):
            data.append([
                str(idx),
                f'#{order.id}',
                order.user.username,
                order.ticket.seminar.title,
                order.ticket.get_ticket_type_display(),
                str(order.quantity),
                f'Rp {order.total_price:,.2f}',
                order.get_status_display(),
                order.order_date.strftime('%Y-%m-%d %H:%M')
            ])

        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        
        elements.append(table)
        doc.build(elements)
        
        buffer.seek(0)
        filename = f"laporan_pesanan_{timezone.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        
        return FileResponse(buffer, as_attachment=True, filename=filename, content_type='application/pdf')

class ExportOrdersXLSXView(APIView):
    """
    Endpoint untuk mengekspor data pesanan ke Excel (XLSX).
    Hanya bisa diakses oleh role admin.
    """
    permission_classes = [IsAdminRole]

    @extend_schema(
        summary='Export Pesanan (XLSX)',
        description='Mengunduh daftar pesanan dalam format Excel (XLSX). Hanya untuk admin.',
        responses={200: bytes}
    )
    def get(self, request):
        orders = Order.objects.all().select_related('user', 'ticket__seminar')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Laporan Pesanan'

        # Headers
        headers = ['No', 'Order ID', 'User', 'Seminar', 'Tipe Tiket', 'Qty', 'Total Harga', 'Status', 'Tanggal']
        ws.append(headers)

        # Style Headers
        header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        header_alignment = Alignment(horizontal='center', vertical='center')

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Data
        for idx, order in enumerate(orders, 1):
            ws.append([
                idx,
                f'#{order.id}',
                order.user.username,
                order.ticket.seminar.title,
                order.ticket.get_ticket_type_display(),
                order.quantity,
                float(order.total_price),
                order.get_status_display(),
                order.order_date.strftime('%Y-%m-%d %H:%M')
            ])

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        # Prepare Response
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        filename = f"laporan_pesanan_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        response = HttpResponse(
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename={filename}'
        
        return response
